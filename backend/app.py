"""
督学管理系统 - Flask 后端主程序
认证 + CRUD API + Bot 调度 + Excel 导入
"""
import os
import sys
import json
import re
import time
import threading
import uuid
from datetime import datetime, timedelta
from functools import wraps
from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS

# Resolve project root (parent of backend/)
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# Fallback to CWD if __file__ resolution is wrong
if not os.path.exists(PROJECT_ROOT):
    PROJECT_ROOT = os.getcwd()
STATIC_DIR = os.path.join(PROJECT_ROOT, 'static')
# Also check backend/static
BACKEND_STATIC = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static')
if os.path.exists(BACKEND_STATIC):
    STATIC_DIR = BACKEND_STATIC
DB_PATH = os.path.join(PROJECT_ROOT, 'backend', 'supervision.db')

app = Flask(__name__)
CORS(app)
import database as db

# ============ Auth Decorators ============

def require_auth(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        if not token:
            return jsonify({'error': '未登录', 'code': 'UNAUTHORIZED'}), 401
        user = db.get_session_user(token)
        if not user:
            return jsonify({'error': '会话已过期', 'code': 'SESSION_EXPIRED'}), 401
        request.user = user
        return f(*args, **kwargs)
    return decorated


def require_admin(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        if not token:
            return jsonify({'error': '未登录', 'code': 'UNAUTHORIZED'}), 401
        user = db.get_session_user(token)
        if not user:
            return jsonify({'error': '会话已过期', 'code': 'SESSION_EXPIRED'}), 401
        request.user = user
        if user.get('role') != 'admin':
            return jsonify({'error': '权限不足', 'code': 'FORBIDDEN'}), 403
        return f(*args, **kwargs)
    return decorated


def require_admin_or_leader(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        if not token:
            return jsonify({'error': '未登录', 'code': 'UNAUTHORIZED'}), 401
        user = db.get_session_user(token)
        if not user:
            return jsonify({'error': '会话已过期', 'code': 'SESSION_EXPIRED'}), 401
        request.user = user
        if user.get('role') not in ('admin', 'leader'):
            return jsonify({'error': '权限不足', 'code': 'FORBIDDEN'}), 403
        return f(*args, **kwargs)
    return decorated


def _get_user_group_ids(user):
    """获取用户可见的分组 ID 列表"""
    if user['role'] == 'admin':
        return None  # None = all groups
    groups = db.get_user_groups(user['user_id'])
    return [g['id'] for g in groups]


def _get_leader_group_ids(user):
    """获取组长的发送权限分组 ID"""
    if user['role'] == 'admin':
        return None
    groups = db.get_leader_groups(user['user_id'])
    return [g['id'] for g in groups]


# ============ Health Check ============

@app.route('/api/health')
def health():
    return jsonify({'status': 'ok', 'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')})


# ============ Auth API ============

@app.route('/api/auth/login', methods=['POST'])
def login():
    data = request.json or {}
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()

    if not username or not password:
        return jsonify({'error': '用户名和密码不能为空'}), 400

    user = db.authenticate(username, password)
    if not user:
        return jsonify({'error': '用户名或密码错误'}), 401

    if not user['is_active']:
        return jsonify({'error': '账号已禁用'}), 403

    token = db.create_session(user['id'])
    group_ids = [g['id'] for g in db.get_user_groups(user['id'])]
    leader_group_ids = [g['id'] for g in db.get_leader_groups(user['id'])] if user['role'] == 'leader' else []
    return jsonify({
        'token': token,
        'user': {
            'id': user['id'],
            'username': user['username'],
            'displayName': user['display_name'],
            'role': user['role'],
            'groupIds': group_ids,
            'leaderGroupIds': leader_group_ids,
        }
    })


@app.route('/api/auth/logout', methods=['POST'])
@require_auth
def logout():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    db.delete_session(token)
    return jsonify({'message': '已登出'})


@app.route('/api/auth/me', methods=['GET'])
@app.route('/api/auth/verify', methods=['GET'])
def me():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    if not token:
        return jsonify({'user': None})
    user = db.get_session_user(token)
    if user:
        group_ids = [g['id'] for g in db.get_user_groups(user['user_id'])]
        leader_group_ids = [g['id'] for g in db.get_leader_groups(user['user_id'])] if user['role'] == 'leader' else []
        return jsonify({'user': {
            'id': user['user_id'],
            'username': user['username'],
            'displayName': user['display_name'],
            'role': user['role'],
            'groupIds': group_ids,
            'leaderGroupIds': leader_group_ids,
        }})
    return jsonify({'user': None})


# ============ Stats API ============

@app.route('/api/stats')
@require_auth
def stats():
    group_ids = _get_user_group_ids(request.user)
    return jsonify(db.get_stats(group_ids))


# ============ Students API ============

@app.route('/api/students', methods=['GET'])
@require_auth
def list_students():
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('pageSize', 50, type=int)
    keyword = request.args.get('keyword', '').strip()
    group_id = request.args.get('groupId', '').strip() or None

    user_groups = _get_user_group_ids(request.user)

    students, total = db.get_students(
        page=page, page_size=page_size, keyword=keyword,
        group_id=group_id, user_groups=user_groups
    )
    return jsonify({'students': students, 'total': total, 'page': page, 'pageSize': page_size})


@app.route('/api/students', methods=['POST'])
@require_auth
def create_student():
    data = request.json or {}
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': '学员姓名不能为空'}), 400

    sid = db.create_student(
        name=name,
        wechat_group_name=data.get('wechat_group_name', '').strip(),
        data_group_id=data.get('data_group_id')
    )
    return jsonify({'id': sid, 'message': '创建成功'}), 201


@app.route('/api/students/<student_id>', methods=['PUT'])
@require_auth
def update_student(student_id):
    data = request.json or {}
    kwargs = {}
    for k in ('name', 'wechat_group_name', 'data_group_id'):
        if k in data:
            kwargs[k] = data[k].strip() if isinstance(data[k], str) else data[k]
    if kwargs:
        db.update_student(student_id, **kwargs)
    return jsonify({'message': '更新成功'})


@app.route('/api/students/<student_id>', methods=['DELETE'])
@require_auth
def delete_student(student_id):
    db.delete_student(student_id)
    return jsonify({'message': '删除成功'})


@app.route('/api/students/<student_id>/records', methods=['GET'])
@require_auth
def student_records(student_id):
    records = db.get_student_records(student_id)
    return jsonify({'records': records})


@app.route('/api/students/<student_id>/records', methods=['POST'])
@require_auth
def add_record(student_id):
    data = request.json or {}
    rid = db.add_student_record(
        student_id=student_id,
        teaching_date=data.get('teaching_date', ''),
        time_period=data.get('time_period', ''),
        course_name=data.get('course_name', ''),
        course_link=data.get('course_link', ''),
        supervision_script=data.get('supervision_script', ''),
        supervision_status=data.get('supervision_status', '课前30分钟发送'),
        wechat_group_name=data.get('wechat_group_name', '')
    )
    return jsonify({'id': rid, 'message': '添加成功'}), 201


@app.route('/api/students/<student_id>/records/<record_id>', methods=['PUT'])
@require_auth
def update_record(student_id, record_id):
    data = request.json or {}
    kwargs = {}
    for k in ('teaching_date', 'time_period', 'course_name', 'course_link',
              'supervision_script', 'supervision_status', 'wechat_group_name'):
        if k in data:
            kwargs[k] = data[k]
    if kwargs:
        db.update_student_record(record_id, **kwargs)
    return jsonify({'message': '更新成功'})


@app.route('/api/students/<student_id>/records/<record_id>', methods=['DELETE'])
@require_auth
def delete_record(student_id, record_id):
    db.delete_student_record(record_id)
    return jsonify({'message': '删除成功'})


# ============ All Records API ============

@app.route('/api/all-records', methods=['GET'])
@require_auth
def all_records():
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('pageSize', 50, type=int)
    date_from = request.args.get('dateFrom', '').strip()
    date_to = request.args.get('dateTo', '').strip()
    group_id = request.args.get('groupId', '').strip() or None
    status = request.args.get('status', '').strip()

    user_groups = _get_user_group_ids(request.user)

    records, total = db.get_all_records(
        date_from=date_from or None, date_to=date_to or None,
        group_id=group_id, status=status or None,
        user_groups=user_groups, page=page, page_size=page_size
    )
    return jsonify({'records': records, 'total': total, 'page': page, 'pageSize': page_size})


# ============ Import API ============

@app.route('/api/import', methods=['POST'])
@require_auth
def import_excel():
    """导入课表数据 — 支持两种模式：
    1. JSON 模式 (Content-Type: application/json)：前端解析 Excel 后发送结构化数据
    2. 文件上传模式 (multipart/form-data)：后端解析 Excel 文件
    """
    content_type = request.content_type or ''

    # ---- JSON 模式：前端已解析 ----
    if 'application/json' in content_type:
        data = request.json or {}
        records = data.get('records', [])
        data_group_id = data.get('dataGroupId')
        wechat_group_name = data.get('wechatGroupName', '')
        source_file = data.get('sourceFile', '')

        if not records:
            return jsonify({'error': '没有可导入的记录'}), 400
        if not data_group_id:
            return jsonify({'error': '请选择导入分组'}), 400

        try:
            result = db.import_records_json(
                records=records,
                data_group_id=data_group_id,
                wechat_group_name=wechat_group_name,
                source_file=source_file,
            )
            return jsonify({
                'success': True,
                'newStudents': result['new_students'],
                'newRecords': result['new_records'],
                'total': len(records),
            })
        except Exception as e:
            import traceback
            traceback.print_exc()
            return jsonify({'error': f'导入失败: {str(e)}'}), 500

    # ---- 文件上传模式：后端解析 ----
    if 'file' not in request.files:
        return jsonify({'error': '未上传文件，请使用 JSON 格式或上传 Excel 文件'}), 400

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': '仅支持 Excel 文件'}), 400

    try:
        from openpyxl import load_workbook
        import tempfile

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=file.filename)
        file.save(tmp.name)
        tmp.close()

        wb = load_workbook(tmp.name, read_only=True, data_only=True)
        ws = wb.active

        rows_data = []
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(c or '').strip() if c else '' for c in row]
                continue
            if i > 5000:
                break

            row_dict = {}
            for j, val in enumerate(row):
                if j < len(headers) and headers[j]:
                    row_dict[headers[j]] = str(val or '').strip() if val is not None else ''
            rows_data.append(row_dict)

        wb.close()
        os.unlink(tmp.name)

        field_map = {
            '学员姓名': 'name', '姓名': 'name', '学员': 'name', 'name': 'name',
            '微信群': 'wechat_group_name', '群名称': 'wechat_group_name', '群名': 'wechat_group_name',
            '分组': 'group_name', '组': 'group_name', 'group': 'group_name',
            '上课日期': 'teaching_date', '日期': 'teaching_date', 'date': 'teaching_date',
            '时间段': 'time_period', '时间': 'time_period', 'time': 'time_period',
            '课程': 'course_name', '课程名称': 'course_name', 'course': 'course_name',
            '课程链接': 'course_link', '链接': 'course_link', 'link': 'course_link',
            '督学话术': 'supervision_script', '话术': 'supervision_script', 'script': 'supervision_script',
            '督学状态': 'supervision_status', '状态': 'supervision_status', 'status': 'supervision_status',
        }

        mapped = []
        for r in rows_data:
            new_row = {}
            for k, v in r.items():
                mapped_key = field_map.get(k, field_map.get(k.lower(), ''))
                if mapped_key:
                    new_row[mapped_key] = v
            if new_row.get('name'):
                mapped.append(new_row)

        imported = db.import_students_and_records(mapped)
        return jsonify({'imported': imported, 'total': len(mapped)})

    except Exception as e:
        return jsonify({'error': f'导入失败: {str(e)}'}), 500


# ============ My Groups API ============

@app.route('/api/my-groups', methods=['GET'])
@require_auth
def my_groups():
    if request.user['role'] == 'admin':
        return jsonify({'groups': db.get_all_groups()})
    return jsonify({'groups': db.get_user_groups(request.user['user_id'])})


# ============ Admin: Users API ============

@app.route('/api/admin/users', methods=['GET'])
@require_admin
def admin_users():
    return jsonify({'users': db.get_all_users()})


@app.route('/api/admin/users', methods=['POST'])
@require_admin
def admin_create_user():
    data = request.json or {}
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()
    display_name = data.get('display_name', '').strip()
    role = data.get('role', 'user')

    if not username or not password:
        return jsonify({'error': '用户名和密码不能为空'}), 400

    pw_hash = db.hash_password(password)
    uid = db.create_user(username, pw_hash, display_name, role)

    # Set groups
    group_ids = data.get('group_ids', [])
    if group_ids:
        db.set_user_groups(uid, group_ids)

    # Set leader groups
    if role == 'leader':
        leader_group_ids = data.get('leader_group_ids', [])
        db.set_leader_groups(uid, leader_group_ids)

    return jsonify({'id': uid, 'message': '创建成功'}), 201


@app.route('/api/admin/users/<user_id>', methods=['PUT'])
@require_admin
def admin_update_user(user_id):
    data = request.json or {}
    kwargs = {}
    for k in ('username', 'display_name', 'role', 'is_active'):
        if k in data:
            kwargs[k] = data[k].strip() if isinstance(data[k], str) else data[k]
    if 'password' in data and data['password'].strip():
        kwargs['password_hash'] = db.hash_password(data['password'].strip())
    if kwargs:
        db.update_user(user_id, **kwargs)

    group_ids = data.get('group_ids')
    if group_ids is not None:
        db.set_user_groups(user_id, group_ids)

    leader_group_ids = data.get('leader_group_ids')
    if leader_group_ids is not None:
        db.set_leader_groups(user_id, leader_group_ids)

    return jsonify({'message': '更新成功'})


@app.route('/api/admin/users/<user_id>', methods=['DELETE'])
@require_admin
def admin_delete_user(user_id):
    if user_id == request.user['user_id']:
        return jsonify({'error': '不能删除自己'}), 400
    db.delete_user(user_id)
    return jsonify({'message': '删除成功'})


# ============ Admin: Groups API ============

@app.route('/api/admin/groups', methods=['GET'])
@require_admin
def admin_groups():
    return jsonify({'groups': db.get_all_groups()})


@app.route('/api/admin/groups', methods=['POST'])
@require_admin
def admin_create_group():
    data = request.json or {}
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': '分组名称不能为空'}), 400
    gid = db.create_group(name, data.get('description', ''))
    return jsonify({'id': gid, 'message': '创建成功'}), 201


@app.route('/api/admin/groups/<group_id>', methods=['PUT'])
@require_admin
def admin_update_group(group_id):
    data = request.json or {}
    kwargs = {}
    for k in ('name', 'description'):
        if k in data:
            kwargs[k] = data[k].strip() if isinstance(data[k], str) else data[k]
    if kwargs:
        db.update_group(group_id, **kwargs)
    return jsonify({'message': '更新成功'})


@app.route('/api/admin/groups/<group_id>', methods=['DELETE'])
@require_admin
def admin_delete_group(group_id):
    db.delete_group(group_id)
    return jsonify({'message': '删除成功'})


# ============ Bot: 桌面自动化督学（任务队列） ============

class BotScheduler:
    """Bot 调度器：管理任务队列，实际发送由 wecom-sender (pywinauto) 完成"""

    def __init__(self):
        self.running = False
        self.thread = None
        self.tasks = []
        self.lock = threading.Lock()
        self._sent_task_ids = set()
        self._sending_task_ids = set()

    def load_tasks(self):
        """从数据库加载今日待发送任务"""
        today = datetime.now().strftime('%Y-%m-%d')
        with db.get_db() as conn:
            rows = conn.execute(
                "SELECT sr.*, s.name as student_name, s.wechat_group_name as student_wx_group "
                "FROM student_records sr "
                "JOIN students s ON sr.student_id = s.id "
                "WHERE sr.teaching_date = ?",
                (today,)
            ).fetchall()

        new_tasks = []
        for r in rows:
            record = dict(r)
            send_time = self._calc_send_time(record)
            if send_time:
                record['_send_time'] = send_time
                record['_task_id'] = str(uuid.uuid4())
                record['_message'] = self._build_message(record)
                new_tasks.append(record)

        with self.lock:
            self.tasks = [t for t in self.tasks
                         if t['_task_id'] in self._sent_task_ids or t['_task_id'] in self._sending_task_ids]
            self.tasks += new_tasks
            self.tasks.sort(key=lambda t: t['_send_time'])

    def _calc_send_time(self, record):
        time_period = record.get('time_period', '').strip()
        if not time_period or '-' not in time_period:
            return None
        parts = time_period.split('-')
        if len(parts) != 2:
            return None
        try:
            start_h, start_m = map(int, parts[0].strip().split(':'))
            end_h, end_m = map(int, parts[1].strip().split(':'))
            status = record.get('supervision_status', '课前30分钟发送')
            now = datetime.now()
            base = now.replace(hour=start_h, minute=start_m, second=0, microsecond=0)
            if '课前30' in status:
                return base - timedelta(minutes=30)
            elif '课前1小时' in status:
                return base - timedelta(hours=1)
            elif '上课中' in status or '上课' in status:
                return base
            elif '课后30' in status:
                return now.replace(hour=end_h, minute=end_m, second=0) + timedelta(minutes=30)
            elif '课后1小时' in status:
                return now.replace(hour=end_h, minute=end_m, second=0) + timedelta(hours=1)
            else:
                return base - timedelta(minutes=30)
        except (ValueError, IndexError):
            return None

    def _build_message(self, record):
        script = record.get('supervision_script', '')
        student_name = record.get('student_name', '')
        course_name = record.get('course_name', '')
        time_period = record.get('time_period', '')
        course_link = record.get('course_link', '')

        if script:
            msg = script
            if time_period and '-' in time_period:
                msg = re.sub(r'XX点XX分|xx点xx分', time_period.split('-')[0].strip(), msg)
            msg = re.sub(r'XX课程|XX课|xx课程|xx课', course_name, msg)
            msg = re.sub(r'XX同学|xx同学', student_name, msg)
            msg = re.sub(r'课程链接:XX|课程链接:xx', course_link, msg)
            return msg
        return f"【督学提醒】{student_name}同学，{course_name}课程即将开始（{time_period}），请准时上课！"

    def _run(self):
        while self.running:
            try:
                self.load_tasks()
            except Exception as e:
                print(f"[Bot Error] {e}")
            time.sleep(30)

    def start(self):
        if self.running:
            return
        self._sent_task_ids.clear()
        self._sending_task_ids.clear()
        self.load_tasks()
        self.running = True
        self.thread = threading.Thread(target=self._run, daemon=True)
        self.thread.start()
        print("[Bot] 已启动，桌面自动化模式（等待 wecom-sender 连接）")

    def stop(self):
        self.running = False
        if self.thread:
            self.thread.join(timeout=5)
            self.thread = None
        print("[Bot] 已停止")

    def get_status(self):
        now = datetime.now()
        pending = 0
        ready = 0
        with self.lock:
            for t in self.tasks:
                tid = t['_task_id']
                if tid not in self._sent_task_ids and tid not in self._sending_task_ids:
                    if t.get('_send_time') and now >= t['_send_time']:
                        ready += 1
                    else:
                        pending += 1
        return {
            'running': self.running,
            'mode': 'desktop_automation',
            'total_tasks': len(self.tasks),
            'pending': pending,
            'ready': ready,
            'sending': len(self._sending_task_ids),
            'sent': len(self._sent_task_ids),
        }

    def claim_tasks(self, limit=5):
        """wecom-sender 认领就绪任务"""
        now = datetime.now()
        claimed = []
        with self.lock:
            for t in self.tasks:
                tid = t['_task_id']
                if tid in self._sent_task_ids or tid in self._sending_task_ids:
                    continue
                if t.get('_send_time') and now >= t['_send_time']:
                    self._sending_task_ids.add(tid)
                    claimed.append({
                        'task_id': tid,
                        'student_name': t.get('student_name', ''),
                        'course_name': t.get('course_name', ''),
                        'group_name': t.get('wechat_group_name', ''),
                        'message': t.get('_message', ''),
                        'send_time': t['_send_time'].strftime('%Y-%m-%d %H:%M:%S'),
                    })
                    if len(claimed) >= limit:
                        break
        return claimed

    def mark_send_result(self, task_id, success, error_msg=''):
        """wecom-sender 汇报发送结果"""
        with self.lock:
            for t in self.tasks:
                if t['_task_id'] == task_id:
                    self._sending_task_ids.discard(task_id)
                    if success:
                        self._sent_task_ids.add(task_id)
                    group_name = t.get('wechat_group_name', '')
                    student_name = t.get('student_name', '')
                    course_name = t.get('course_name', '')
                    db.add_send_log(
                        task_id=task_id, group_name=group_name,
                        student_name=student_name, course_name=course_name,
                        success=success, error_msg=error_msg, send_mode='desktop_automation'
                    )
                    return True
        return False

    def clear_completed(self):
        with self.lock:
            self.tasks = [t for t in self.tasks if t['_task_id'] not in self._sent_task_ids]
            self._sent_task_ids.clear()
            self._sending_task_ids.clear()


bot = BotScheduler()


# ============ Bot API（桌面自动化模式） ============

@app.route('/api/bot/status')
@require_auth
def bot_status():
    return jsonify(bot.get_status())


@app.route('/api/bot/start', methods=['POST'])
@require_admin_or_leader
def bot_start():
    bot.start()
    return jsonify({'message': 'Bot 已启动，等待桌面自动化发送器连接', 'status': bot.get_status()})


@app.route('/api/bot/stop', methods=['POST'])
@require_admin_or_leader
def bot_stop():
    bot.stop()
    return jsonify({'message': 'Bot 已停止', 'status': bot.get_status()})


@app.route('/api/bot/claim-tasks', methods=['POST'])
@require_auth
def bot_claim_tasks():
    """wecom-sender 调用：获取就绪任务"""
    data = request.json or {}
    limit = data.get('limit', 3)
    tasks = bot.claim_tasks(limit=limit)
    return jsonify({'tasks': tasks})


@app.route('/api/bot/send-result', methods=['POST'])
@require_auth
def bot_send_result():
    """wecom-sender 调用：汇报发送结果"""
    data = request.json or {}
    task_id = data.get('task_id', '')
    success = data.get('success', False)
    error_msg = data.get('error_msg', '')
    if not task_id:
        return jsonify({'error': '缺少 task_id'}), 400
    bot.mark_send_result(task_id, success, error_msg)
    return jsonify({'message': 'ok'})


@app.route('/api/bot/logs')
@require_auth
def bot_logs():
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('pageSize', 50, type=int)
    success_filter = request.args.get('success', type=int)
    logs, total = db.get_send_logs(page=page, page_size=page_size, success_filter=success_filter)
    return jsonify({'logs': logs, 'total': total, 'page': page, 'pageSize': page_size})


@app.route('/api/bot/clear-completed', methods=['POST'])
@require_admin_or_leader
def bot_clear():
    bot.clear_completed()
    return jsonify({'message': '已清空已完成任务'})


# ============ Frontend SPA Fallback ============

@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def serve_frontend(path):
    print(f"[serve_frontend] path={path!r}, STATIC_DIR={STATIC_DIR}, exists={os.path.exists(STATIC_DIR)}")
    if path and os.path.exists(os.path.join(STATIC_DIR, path)):
        return send_from_directory(STATIC_DIR, path)
    return send_from_directory(STATIC_DIR, 'index.html')


# ============ Entry Point ============

if __name__ == '__main__':
    db.init_db()
    print("=" * 50)
    print("  督学管理系统 - Flask Backend")
    print("  http://localhost:8080")
    print("  默认账号: admin / admin123")
    print(f"  Static folder: {app.static_folder}")
    print("=" * 50)
    port = int(os.environ.get('PORT', 8888))
    app.run(host='0.0.0.0', port=port, debug=False, threaded=True)
